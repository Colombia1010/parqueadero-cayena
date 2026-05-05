"""
blueprints/admin.py
────────────────────
Rutas protegidas del panel de administración:
- Vista principal con registros y vehículos autorizados
- CRUD de vehículos autorizados (uno a uno y carga masiva Excel)
- Exportación a Excel
- Descarga de documentos
"""

import io
import os
from functools import wraps

from flask import (
    Blueprint, request, render_template, redirect,
    session, send_from_directory, send_file, current_app
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from app.managers.usuario_manager  import UsuarioManager
from app.managers.vehiculo_manager import VehiculoManager
from app.services.usuario_servicio import UsuarioServicio
from app.utils.validaciones        import validar_torre, validar_apto, validar_placa

admin_bp = Blueprint("admin", __name__)


# ── Decorador de sesión ────────────────────────────────────────────────────────

def requiere_admin(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("admin"):
            return redirect("/login")
        return f(*args, **kwargs)
    return wrapper


# ── Panel principal ────────────────────────────────────────────────────────────

@admin_bp.route("/admin")
@requiere_admin
def admin():
    servicio  = UsuarioServicio()
    usuarios  = servicio.obtener_usuarios_con_estado()
    vehiculos = VehiculoManager().obtener_todos()
    return render_template("admin.html", usuarios=usuarios, vehiculos=vehiculos)


# ── Agregar vehículo uno a uno ─────────────────────────────────────────────────

@admin_bp.route("/admin/agregar_vehiculo", methods=["POST"])
@requiere_admin
def agregar_vehiculo():
    placa = request.form.get("placa", "").strip().upper()

    ok, msg = validar_placa(placa)
    if not ok:
        current_app.logger.warning("⚠️ Placa inválida al agregar: %s — %s", placa, msg)
        return redirect("/admin")

    torre = request.form.get("torre", "").strip()
    apto  = request.form.get("apto", "").strip()

    ok_torre, _ = validar_torre(torre)
    ok_apto,  _ = validar_apto(apto)

    if not ok_torre or not ok_apto:
        return redirect("/admin")

    datos = {
        "placa":            placa,
        "tipo":             request.form.get("tipo", "").strip(),
        "torre":            torre,
        "apto":             apto,
        "nombre_residente": request.form.get("nombre_residente", "").strip(),
    }

    ok, msg = VehiculoManager().agregar(datos)

    if ok:
        current_app.logger.info(
            "✅ Vehículo autorizado: %s | %s | Torre %s Apto %s",
            placa, datos["nombre_residente"], torre, apto
        )
    else:
        current_app.logger.warning("⚠️ No se pudo autorizar placa %s: %s", placa, msg)

    return redirect("/admin")


# ── Carga masiva desde Excel ───────────────────────────────────────────────────

@admin_bp.route("/admin/cargar_excel", methods=["POST"])
@requiere_admin
def cargar_excel():
    """
    Procesa el Excel del conjunto Cayena (hoja CONSOLIDADO AUTOS Y MOTOS)
    y extrae todas las placas de autos, motos y bicicletas eléctricas.
    Agrega solo las que no estén ya registradas — se puede subir el
    mismo archivo varias veces sin generar duplicados.
    """
    archivo = request.files.get("archivo_excel")

    if not archivo or archivo.filename == "":
        current_app.logger.warning("⚠️ Carga Excel sin archivo adjunto")
        return redirect("/admin")

    ext = archivo.filename.rsplit(".", 1)[-1].lower()
    if ext not in {"xlsx", "xls"}:
        current_app.logger.warning("⚠️ Formato inválido en carga Excel: %s", archivo.filename)
        return redirect("/admin")

    try:
        import pandas as pd

        df = pd.read_excel(
            archivo,
            sheet_name="CONSOLIDADO AUTOS Y MOTOS",
            dtype=str
        )

        manager   = VehiculoManager()
        agregados = 0
        omitidos  = 0

        for _, fila in df.iterrows():
            # Torre y apto: quitar decimales que pandas agrega (5.0 → 5)
            torre  = str(fila.get("TO",   "") or "").strip().split(".")[0]
            apto   = str(fila.get("APTO", "") or "").strip().split(".")[0]
            nombre = str(fila.get("NOMBRE TARJETA DE PROPIEDAD", "") or "").strip()

            if not torre or not apto or torre == "nan" or apto == "nan":
                continue

            # ── Auto ──────────────────────────────────────────────────────────
            placa_auto = str(fila.get("PLACA", "") or "").strip().upper()
            if placa_auto and placa_auto != "NAN" and len(placa_auto) <= 6:
                ok, _ = manager.agregar({
                    "placa":            placa_auto,
                    "tipo":             "Carro",
                    "torre":            torre,
                    "apto":             apto,
                    "nombre_residente": nombre,
                })
                agregados += 1 if ok else 0
                omitidos  += 0 if ok else 1

            # ── Moto ──────────────────────────────────────────────────────────
            placa_moto = str(fila.get("PLACA.1", "") or "").strip().upper()
            if placa_moto and placa_moto != "NAN" and len(placa_moto) <= 6:
                ok, _ = manager.agregar({
                    "placa":            placa_moto,
                    "tipo":             "Moto",
                    "torre":            torre,
                    "apto":             apto,
                    "nombre_residente": nombre,
                })
                agregados += 1 if ok else 0
                omitidos  += 0 if ok else 1

            # ── Bicicleta eléctrica ────────────────────────────────────────────
            bici = str(fila.get("BICI ELECTRICA", "") or "").strip().upper()
            if bici and bici != "NAN":
                ok, _ = manager.agregar({
                    "placa":            bici[:6],
                    "tipo":             "Bicicleta Electrica",
                    "torre":            torre,
                    "apto":             apto,
                    "nombre_residente": nombre,
                })
                agregados += 1 if ok else 0
                omitidos  += 0 if ok else 1

        current_app.logger.info(
            "📥 Carga Excel — nuevos: %d | ya existían: %d", agregados, omitidos
        )

    except Exception as e:
        current_app.logger.error("❌ Error procesando Excel: %s", str(e), exc_info=True)

    return redirect("/admin")


# ── Eliminar vehículo ──────────────────────────────────────────────────────────

@admin_bp.route("/admin/eliminar_vehiculo/<placa>", methods=["POST"])
@requiere_admin
def eliminar_vehiculo(placa):
    ok, msg = VehiculoManager().eliminar(placa)
    if ok:
        current_app.logger.info("🗑 Vehículo eliminado: %s", placa.upper())
    else:
        current_app.logger.warning("⚠️ No se pudo eliminar placa %s: %s", placa, msg)
    return redirect("/admin")


# ── Exportar Excel de registros ────────────────────────────────────────────────

@admin_bp.route("/exportar_excel")
@requiere_admin
def exportar_excel():
    servicio = UsuarioServicio()
    usuarios = servicio.obtener_usuarios_con_estado()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Cayena"

    ws.merge_cells("A1:Q3")
    try:
        ruta_logo = os.path.join(os.getcwd(), "static", "logo.png")
        if os.path.exists(ruta_logo):
            img        = XLImage(ruta_logo)
            img.width  = 900
            img.height = 150
            ws.add_image(img, "A1")
    except Exception as e:
        current_app.logger.warning("⚠️ No se pudo cargar el logo en Excel: %s", str(e))

    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 60
    ws.row_dimensions[3].height = 40

    ws.merge_cells("A4:Q4")
    ws["A4"]           = "REPORTE PARQUEADERO - CONJUNTO CAYENA"
    ws["A4"].font      = Font(size=14, bold=True)
    ws["A4"].alignment = Alignment(horizontal="center")

    headers = [
        "Cédula", "Nombre", "Torre", "Apto",
        "Nombre Tarjeta", "Celular", "Correo",
        "Propietario", "Mail Propietario",
        "Celular 2", "Arrendatario",
        "Vehículo", "Placa", "Marca", "Modelo", "Color", "Estado"
    ]

    ws.append([])
    ws.append(headers)

    header_row  = ws.max_row
    fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    font_header = Font(bold=True, color="FFFFFF")

    for cell in ws[header_row]:
        cell.fill      = fill_header
        cell.font      = font_header
        cell.alignment = Alignment(horizontal="center")

    for u in usuarios:
        ws.append([
            u.get("cedula"),              u.get("nombre"),
            u.get("torre"),               u.get("apto"),
            u.get("nombre_propiedad"),    u.get("celular"),
            u.get("correo"),              u.get("nombre_propietario"),
            u.get("mail_propietario"),    u.get("celular1"),
            u.get("nombre_arrendatario"), u.get("vehiculo"),
            u.get("placa"),               u.get("marca"),
            u.get("modelo"),              u.get("color"),
            u.get("estado"),
        ])

    fill_gris = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row), start=1):
        if i % 2 == 0:
            for cell in row:
                cell.fill = fill_gris

    ws.sheet_view.showGridLines = False

    for i, col in enumerate(ws.columns, 1):
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value), default=0
        )
        ws.column_dimensions[get_column_letter(i)].width = max_len + 3

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    current_app.logger.info("📥 Excel exportado — %d registros", len(usuarios))

    return send_file(
        stream,
        as_attachment=True,
        download_name="reporte_cayena.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ── Descarga de documentos ─────────────────────────────────────────────────────

@admin_bp.route("/uploads/<cedula>/<filename>")
@requiere_admin
def descargar_archivo(cedula, filename):
    return send_from_directory(
        os.path.join(current_app.config["UPLOAD_FOLDER"], cedula),
        filename
    )
