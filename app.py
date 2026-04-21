from flask import Flask, request, render_template, redirect, session, send_from_directory, send_file
from werkzeug.utils import secure_filename
from usuarios import UsuarioManager
from servicios import UsuarioServicio
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from flask import send_file
import io
import os

app = Flask(__name__)
app.secret_key = "clave_secreta"

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

usuario_manager = UsuarioManager()
usuario_servicio = UsuarioServicio()

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}


# ---------------------------
# VALIDAR ARCHIVOS
# ---------------------------
def archivo_permitido(nombre):
    return '.' in nombre and nombre.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ---------------------------
# HOME
# ---------------------------
@app.route('/')
def inicio():
    return render_template('formulario.html')


# ---------------------------
# REGISTRO
# ---------------------------
@app.route('/registro', methods=['POST'])
def registro():
    try:
        # -------------------------
        # DATOS
        # -------------------------
        cedula = request.form['cedula']
        nombre = request.form['nombre']
        torre = request.form['torre']
        apto = request.form['apto']
        nombre_propiedad = request.form['nombre_propiedad']
        celular = request.form['celular']
        correo = request.form['correo']
        nombre_propietario = request.form['nombre_propietario']
        mail_propietario = request.form['mail_propietario']
        celular1 = request.form.get('celular1')
        nombre_arrendatario = request.form.get('nombre_arrendatario')
        vehiculo = request.form['vehiculo']
        placa = request.form['placa']
        marca = request.form['marca']
        modelo = request.form['modelo']
        color = request.form['color']

        # -------------------------
        # VALIDACIONES
        # -------------------------
        if len(torre) > 2:
            return "❌ Torre inválida"

        if len(apto) > 4:
            return "❌ Apto inválido"

        if not celular.isdigit() or len(celular) != 10:
            return "❌ Celular inválido"

        if len(placa) > 6:
            return "❌ Placa inválida"

        # -------------------------
        # VALIDAR DUPLICADO
        # -------------------------
        if usuario_manager.usuario_existe(cedula):
            return "⚠️ Usuario ya registrado (puedes actualizar documentos)"

        # -------------------------
        # CARPETA USUARIO
        # -------------------------
        ruta_usuario = os.path.join(UPLOAD_FOLDER, cedula)
        os.makedirs(ruta_usuario, exist_ok=True)

        # -------------------------
        # GUARDAR ARCHIVOS
        # -------------------------
        archivos_campos = ['cedula_doc', 'soat', 'tarjeta']

        for campo in archivos_campos:
            archivo = request.files.get(campo)

            if archivo and archivo.filename != "":
                if archivo_permitido(archivo.filename):
                    nombre_seguro = secure_filename(f"{campo}_{archivo.filename}")
                    archivo.save(os.path.join(ruta_usuario, nombre_seguro))
                else:
                    return f"Archivo inválido: {campo}"

        # -------------------------
        # CREAR OBJETO USUARIO
        # -------------------------
        datos = {
            "cedula": cedula,
            "nombre": nombre,
            "torre": torre,
            "apto": apto,
            "nombre_propiedad": nombre_propiedad,
            "celular": celular,
            "correo": correo,
            "nombre_propietario": nombre_propietario,
            "mail_propietario": mail_propietario,
            "celular1": celular1,
            "nombre_arrendatario": nombre_arrendatario,
            "vehiculo": vehiculo,
            "placa": placa,
            "marca": marca,
            "modelo": modelo,
            "color": color
        }

        # -------------------------
        # GUARDAR
        # -------------------------
        usuario_manager.guardar_usuario(datos)

        return "✅ Registro exitoso"

    except Exception as e:
        print("ERROR:", e)
        return "❌ Error interno del servidor"


# ---------------------------
# ADMIN
# ---------------------------
@app.route('/admin')
def admin():
    if not session.get('admin'):
        return redirect('/login')

    usuarios = usuario_servicio.obtener_usuarios_con_estado()
    return render_template('admin.html', usuarios=usuarios)


# ---------------------------
# EXPORTAR EXCEL PRO
# ---------------------------
@app.route('/exportar_excel')
def exportar_excel():
    usuarios = usuario_servicio.obtener_usuarios_con_estado()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Cayena"

    # -------------------------
    # HEADER TIPO BANNER
    # -------------------------
    ws.merge_cells('A1:Q3')

    try:
        ruta_logo = os.path.join(os.getcwd(), "static", "logo.png")

        if os.path.exists(ruta_logo):
            img = Image(ruta_logo)
            img.width = 900   # 👈 ancho grande tipo banner
            img.height = 150  # 👈 alto tipo cabecera

            ws.add_image(img, "A1")

        else:
            print("⚠️ Logo no encontrado")

    except Exception as e:
        print("ERROR LOGO:", e)

    # Ajustar altura
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 60
    ws.row_dimensions[3].height = 40

    # -------------------------
    # TÍTULO
    # -------------------------
    ws.merge_cells('A4:Q4')
    ws['A4'] = "REPORTE PARQUEADERO - CONJUNTO CAYENA"
    ws['A4'].font = Font(size=14, bold=True)
    ws['A4'].alignment = Alignment(horizontal="center")

    # -------------------------
    # HEADERS
    # -------------------------
    headers = [
        "Cédula", "Nombre", "Torre", "Apto",
        "Nombre Tarjeta", "Celular", "Correo",
        "Propietario", "Mail Propietario",
        "Celular 2", "Arrendatario",
        "Vehículo", "Placa", "Marca", "Modelo", "Color",
        "Estado"
    ]

    ws.append([])
    ws.append(headers)

    header_row = ws.max_row

    fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")

    for cell in ws[header_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    # -------------------------
    # DATOS
    # -------------------------
    for u in usuarios:
        ws.append([
            u.get("cedula"),
            u.get("nombre"),
            u.get("torre"),
            u.get("apto"),
            u.get("nombre_propiedad"),
            u.get("celular"),
            u.get("correo"),
            u.get("nombre_propietario"),
            u.get("mail_propietario"),
            u.get("celular1"),
            u.get("nombre_arrendatario"),
            u.get("vehiculo"),
            u.get("placa"),
            u.get("marca"),
            u.get("modelo"),
            u.get("color"),
            u.get("estado")
        ])

    # -------------------------
    # FILAS TIPO TABLA (alternadas)
    # -------------------------
    fill_gris = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    for i, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=ws.max_row), start=1):
        if i % 2 == 0:
            for cell in row:
                cell.fill = fill_gris

    # -------------------------
    # SIN GRID (PRO)
    # -------------------------
    ws.sheet_view.showGridLines = False

    # -------------------------
    # AUTO AJUSTE
    # -------------------------
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(i)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column].width = max_length + 3

    # -------------------------
    # EXPORTAR
    # -------------------------
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name="reporte_cayena_pro.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ---------------------------
# LOGIN
# ---------------------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':

        if request.form['usuario'] == "admin" and request.form['password'] == "admin":
            session['admin'] = True
            return redirect('/admin')

        return "❌ Credenciales incorrectas"

    return render_template('login.html')


# ---------------------------
# DESCARGAR ARCHIVOS
# ---------------------------
@app.route('/uploads/<cedula>/<filename>')
def descargar_archivo(cedula, filename):
    return send_from_directory(f'uploads/{cedula}', filename)


# ---------------------------
# RUN
# ---------------------------
if __name__ == '__main__':
    app.run(debug=True)